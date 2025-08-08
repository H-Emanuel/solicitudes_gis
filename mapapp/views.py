import json
from django.shortcuts import render, redirect, get_object_or_404
from django.core import serializers
from .models import Punto, Pregunta, Respuesta, Opcion, Formulario, ImagenRespuesta, Departamento, FormularioRespondidoIP
import logging
from django.contrib import messages
from .forms import FormularioForm, FormularioAparienciaForm
from django.http import JsonResponse
from django.db import IntegrityError
from django.db import models
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
from io import BytesIO
import zipfile
from xml.etree.ElementTree import Element, SubElement, tostring
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from .forms_login import LoginForm
from functools import wraps
from hashids import Hashids
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required


import json

from .models import Formulario, Pregunta, Opcion

HASHIDS_SALT = 'cambia_esto_por_un_salt_secreto'
hashids = Hashids(salt=HASHIDS_SALT, min_length=8)

logger = logging.getLogger(__name__)



def decode_hashid_or_404(hashid):
    ids = hashids.decode(hashid)
    if not ids:
        from django.http import Http404
        raise Http404('ID inválida')
    return ids[0]

def guardar_punto(request, formulario_hashid):
    formulario_id = decode_hashid_or_404(formulario_hashid)
    formulario = get_object_or_404(Formulario, id=formulario_id)
    puntos = Punto.objects.filter(respuesta__pregunta__formulario=formulario).distinct().order_by('id')
    puntos_list = [{'latitud': p.latitud, 'longitud': p.longitud} for p in puntos]
    puntos_json = json.dumps(puntos_list)
    preguntas = formulario.preguntas.all().order_by('id')

    for pregunta in preguntas:
        if pregunta.tipo == 'opcion_multiple':
            pregunta.opciones = Opcion.objects.filter(pregunta=pregunta)
        elif pregunta.tipo == 'valoracion':
            pregunta.rango = list(range(pregunta.rango_minimo, pregunta.rango_maximo + 1))

    imagen_url = formulario.imagen.url if formulario.imagen else None

    estilos_header = {
        'titulo': {
            'bold': formulario.titulo_bold,
            'italic': formulario.titulo_italic,
            'underline': formulario.titulo_underline,
            'color': formulario.titulo_color,
            'font': formulario.titulo_font,
            'align': formulario.titulo_align
        },
        'subtitulo': {
            'bold': formulario.subtitulo_bold,
            'italic': formulario.subtitulo_italic,
            'underline': formulario.subtitulo_underline,
            'color': formulario.subtitulo_color,
            'font': formulario.subtitulo_font,
            'align': formulario.subtitulo_align,
            'fontSize': formulario.subtitulo_fontSize,
        }
    }

    print(f"DEBUG: Valor de formulario.subtitulo desde la DB: {formulario.subtitulo}")
    print(f"DEBUG: Valor de formulario.subtitulo_fontSize desde la DB: {formulario.subtitulo_fontSize}")
    print(f"DEBUG: Valor de estilos_header['subtitulo']['fontSize'] enviado a la plantilla: {estilos_header['subtitulo']['fontSize']}")

    tema_color = formulario.tema_color or '#e8e8e8'

    # --- NUEVO: Control de reenvío de formulario ---
    session_key = f'formulario_respondido_{formulario_id}'
    ya_respondido = request.session.get(session_key, False)
    mostrar_mensaje_ya_respondido = False
    # Obtener IP del usuario
    def get_client_ip(request):
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip
    ip_usuario = get_client_ip(request)
    # Detectar si es admin Django o usuario registrado
    es_usuario_registrado = False
    if request.user.is_authenticated:
        es_usuario_registrado = True
    elif request.session.get('usuario_id'):
        try:
            usuario = User.objects.get(id=request.session['usuario_id'])
            es_usuario_registrado = True
        except User.DoesNotExist:
            pass
    # Consultar si ya respondió por IP (para todos)
    ip_respondido = FormularioRespondidoIP.objects.filter(formulario=formulario, ip=ip_usuario).exists()

    # Solo restringir si ya respondió por IP, la opción está desactivada y NO es usuario registrado/admin
    if not formulario.permitir_volver_a_contestar and ip_respondido and not es_usuario_registrado:
        preguntas = []  # No mostrar preguntas
        mostrar_mensaje_ya_respondido = True

    if request.method == 'POST':
        # Buscar si hay una pregunta tipo 'mapa' en el formulario
        pregunta_mapa = preguntas.filter(tipo='mapa').first() if preguntas else None
        lat = None
        lng = None
        if pregunta_mapa:
            mapa_tipo = pregunta_mapa.mapa_tipo or 'punto'
            lat = request.POST.get(f'respuesta_{pregunta_mapa.id}_lat')
            lng = request.POST.get(f'respuesta_{pregunta_mapa.id}_lng')
            # Validar solo si la pregunta es obligatoria
            if pregunta_mapa.obligatorio:
                if mapa_tipo == 'punto':
                    if not lat or not lng:
                        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                            return JsonResponse({'success': False, 'error': 'Debes seleccionar una ubicación en el mapa para esta pregunta obligatoria.'})
                        messages.error(request, 'Debes seleccionar una ubicación en el mapa para esta pregunta obligatoria.')
                        return redirect(request.path)
                elif mapa_tipo in ['linea', 'poligono']:
                    if not lat:
                        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                            return JsonResponse({'success': False, 'error': 'Debes dibujar la geometría en el mapa para esta pregunta obligatoria.'})
                        messages.error(request, 'Debes dibujar la geometría en el mapa para esta pregunta obligatoria.')
                        return redirect(request.path)
        else:
            lat = request.POST.get('lat')
            lng = request.POST.get('lng')
            # Si no hay pregunta tipo mapa, no exigir lat/lng (no validar aquí)
        portada = request.FILES.get('portada')

        # No guardar nada si estamos en la vista de agregar_pregunta.html (solo preview)
        if request.resolver_match and request.resolver_match.view_name == 'agregar_pregunta':
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'No se puede guardar en modo vista previa.'})
            return redirect('agregar_pregunta', formulario_id=formulario.id)

        # Si no se permite volver a contestar y ya respondió, bloquear POST (excepto usuario autenticado)
        if not formulario.permitir_volver_a_contestar and ya_respondido and not es_usuario_registrado:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Ya has respondido este formulario.'})
            messages.error(request, 'Ya has respondido este formulario.')
            return redirect(request.path)

        # --- CORREGIDO: Siempre crear un punto, aunque no haya pregunta tipo mapa ---
        if pregunta_mapa:
            mapa_tipo = pregunta_mapa.mapa_tipo or 'punto'
            if mapa_tipo == 'punto':
                punto = Punto.objects.create(latitud=lat, longitud=lng)
            else:
                # Para línea/polígono, el punto es solo un placeholder
                punto = Punto.objects.create(latitud=0.0, longitud=0.0)
        else:
            punto = Punto.objects.create(latitud=0.0, longitud=0.0)

        # Guardar respuestas asociadas a ese punto
        for pregunta in preguntas:
            campo_nombre = f'respuesta_{pregunta.id}'
            if pregunta.tipo == 'foto':
                if campo_nombre in request.FILES:
                    foto = request.FILES[campo_nombre]
                    imagen_respuesta = ImagenRespuesta(punto=punto, imagen=foto)
                    imagen_respuesta.save()
                    respuesta = Respuesta(
                        pregunta=pregunta,
                        punto=punto,
                        respuesta=f"Imagen: {foto.name}",
                        imagen_respuesta=imagen_respuesta
                    )
                    respuesta.save()
            elif pregunta.tipo == 'opcion_multiple':
                if pregunta.permitir_multiple:
                    respuestas = request.POST.getlist(campo_nombre)
                    for respuesta_texto in respuestas:
                        if respuesta_texto:
                            Respuesta.objects.create(
                                pregunta=pregunta,
                                punto=punto,
                                respuesta=respuesta_texto
                            )
                else:
                    respuesta_texto = request.POST.get(campo_nombre)
                    if respuesta_texto:
                        Respuesta.objects.create(
                            pregunta=pregunta,
                            punto=punto,
                            respuesta=respuesta_texto
                        )
            elif pregunta.tipo == 'mapa':
                # Guardar la geometría según el tipo de la pregunta
                from .models import Geometria
                mapa_tipo = pregunta.mapa_tipo or 'punto'
                lat_field = f'respuesta_{pregunta.id}_lat'
                lng_field = f'respuesta_{pregunta.id}_lng'
                lat_val = request.POST.get(lat_field)
                lng_val = request.POST.get(lng_field)
                try:
                    if mapa_tipo == 'punto':
                        if lat_val and lng_val:
                            coords = f'[{lat_val},{lng_val}]'
                            Geometria.objects.create(
                                tipo='punto',
                                coordenadas=coords,
                                pregunta=pregunta,
                                punto=punto
                            )
                    elif mapa_tipo in ['linea', 'poligono']:
                        if lat_val:
                            Geometria.objects.create(
                                tipo=mapa_tipo,
                                coordenadas=lat_val,
                                pregunta=pregunta,
                                punto=punto
                            )
                except Exception as e:
                    print(f"[ERROR AL GUARDAR GEOMETRIA] {e}")
                    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'error': f'Error al guardar geometría: {e}'})
                    messages.error(request, f'Error al guardar geometría: {e}')
                    return redirect(request.path)
                continue
            else:
                respuesta_texto = request.POST.get(campo_nombre)
                if respuesta_texto:
                    Respuesta.objects.create(
                        pregunta=pregunta,
                        punto=punto,
                        respuesta=respuesta_texto
                    )
        # Guardar en sesión que ya respondió
        request.session[session_key] = True
        request.session.modified = True
        # Guardar la IP en la base de datos (si no existe para este formulario)
        if not FormularioRespondidoIP.objects.filter(formulario=formulario, ip=ip_usuario).exists():
            FormularioRespondidoIP.objects.create(formulario=formulario, ip=ip_usuario)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'redirect_url': reverse('exito_guardado', kwargs={'formulario_hashid': hashids.encode(formulario.id)})})
        # Si no es AJAX, solo mostrar mensaje de éxito sin redirigir
        messages.success(request, '¡Se han guardado las respuestas!')
        return render(request, 'guardar_punto.html', {
            'puntos_json': puntos_json,
            'preguntas': preguntas, 
            'imagen_url': imagen_url,
            'formulario': formulario,
            'estilos_header': estilos_header,
            'tema_color': tema_color,
            'modo_visualizacion': False,
            'mostrar_mensaje_ya_respondido': mostrar_mensaje_ya_respondido
        })

    return render(request, 'guardar_punto.html', {
        'puntos_json': puntos_json,
        'preguntas': preguntas, 
        'imagen_url': imagen_url,
        'formulario': formulario,
        'estilos_header': estilos_header,
        'tema_color': tema_color,
        'modo_visualizacion': False,
        'mostrar_mensaje_ya_respondido': mostrar_mensaje_ya_respondido
    })

def exito(request):
    return render(request, 'exito.html')

def exito_guardado(request, formulario_hashid):
    formulario_id = decode_hashid_or_404(formulario_hashid)
    formulario = None
    try:
        formulario = Formulario.objects.get(id=formulario_id)
    except Formulario.DoesNotExist:
        formulario = None
    return render(request, 'exito_guardado.html', {
        'formulario': formulario,
        'titulo': formulario.titulo if formulario else 'Encuesta',
        'tema_color': formulario.tema_color if formulario and formulario.tema_color else '#b7dfc4',
        'imagen_url': formulario.imagen.url if formulario and formulario.imagen else None
    })

def ver_guardado(request, formulario_hashid):
    formulario_id = decode_hashid_or_404(formulario_hashid)
    formulario = get_object_or_404(Formulario, id=formulario_id)
    puntos = Punto.objects.filter(respuesta__pregunta__formulario=formulario).distinct().order_by('id')
    preguntas = formulario.preguntas.all()
    puntos_con_respuestas = []
    puntos_mapa = []
    from .models import Geometria

    for punto in puntos:
        punto_data = {'punto': punto, 'respuestas': {}, 'imagenes': []}
        # Buscar geometría asociada a este punto
        geometria = Geometria.objects.filter(punto=punto).first()
        if geometria:
            try:
                coords = json.loads(geometria.coordenadas)
            except Exception:
                coords = None
            if geometria.tipo == 'punto' and isinstance(coords, list) and len(coords) == 2:
                lat, lng = coords
                punto_data['punto'].latitud = lat
                punto_data['punto'].longitud = lng
                puntos_mapa.append({
                    'id': punto.id,
                    'tipo': 'punto',
                    'latitud': lat,
                    'longitud': lng,
                    'geometria': coords
                })
            elif geometria.tipo in ['linea', 'poligono'] and isinstance(coords, list) and len(coords) > 0:
                lat, lng = coords[0]
                punto_data['punto'].latitud = lat
                punto_data['punto'].longitud = lng
                puntos_mapa.append({
                    'id': punto.id,
                    'tipo': geometria.tipo,
                    'latitud': lat,
                    'longitud': lng,
                    'geometria': coords
                })
            else:
                puntos_mapa.append({
                    'id': punto.id,
                    'tipo': 'desconocido',
                    'latitud': punto.latitud,
                    'longitud': punto.longitud,
                    'geometria': None
                })
        else:
            puntos_mapa.append({
                'id': punto.id,
                'tipo': 'punto',
                'latitud': punto.latitud,
                'longitud': punto.longitud,
                'geometria': None
            })
        for pregunta in preguntas:
            try:
                respuestas = Respuesta.objects.filter(punto=punto, pregunta=pregunta)
                if pregunta.tipo == 'foto' and respuestas.exists():
                    for respuesta in respuestas:
                        if respuesta.imagen_respuesta:
                            punto_data['respuestas'][pregunta.id] = {
                                'tipo': 'foto',
                                'url': respuesta.imagen_respuesta.imagen.url,
                                'nombre': respuesta.imagen_respuesta.imagen.name
                            }
                else:
                    if respuestas.exists():
                        valores = [respuesta.respuesta for respuesta in respuestas]
                        punto_data['respuestas'][pregunta.id] = {
                            'tipo': 'texto',
                            'valor': ', '.join(valores)
                        }
                    else:
                        punto_data['respuestas'][pregunta.id] = {
                            'tipo': 'texto',
                            'valor': "No hay respuesta"
                        }
            except Exception as e:
                punto_data['respuestas'][pregunta.id] = {
                    'tipo': 'texto',
                    'valor': "Error al cargar respuesta"
                }
        puntos_con_respuestas.append(punto_data)

    puntos_mapa_json = json.dumps(puntos_mapa, indent=2) if puntos_mapa else '[]'
    return render(request, 'ver_guardado.html', {
        'puntos_con_respuestas': puntos_con_respuestas,
        'preguntas': preguntas,
        'formulario': formulario,
        'puntos_mapa': puntos_mapa_json,
    })

def agregar_pregunta(request, formulario_hashid):
    formulario_id = decode_hashid_or_404(formulario_hashid)
    formulario = get_object_or_404(Formulario, id=formulario_id)
    apariencia_form = FormularioAparienciaForm(request.POST or None, instance=formulario)
    color_temas = [
        '#b7e3dd', '#ffc6c6', '#aee9f7', '#f3f3f3', '#6d6d6d',
        '#2e5d4d', '#2d7be5', '#b36a36', '#a34d4d', '#7a5c52'
    ]
    departamentos = Departamento.objects.all()

    # --- NUEVO: Forzar valores por defecto en el backend ---
    if not formulario.titulo or formulario.titulo == 'none':
        formulario.titulo = 'titulo'
    if not formulario.subtitulo or formulario.subtitulo == 'none':
        formulario.subtitulo = 'subtitulo'

    if request.method == 'POST':
        if 'actualizar_formulario' in request.POST:
            # Actualizar información del formulario
            formulario.titulo = request.POST.get('titulo')
            formulario.subtitulo = request.POST.get('subtitulo')
            formulario.tema_color = request.POST.get('tema_color', formulario.tema_color)
            # Guardar texto del botón enviar si viene en el POST
            texto_boton_enviar = request.POST.get('texto_boton_enviar')
            if texto_boton_enviar is not None:
                formulario.texto_boton_enviar = texto_boton_enviar
            if 'imagen' in request.FILES:
                # Eliminar la imagen anterior si existe
                if formulario.imagen:
                    formulario.imagen.delete()
                # Guardar la nueva imagen
                formulario.imagen = request.FILES['imagen']
            # --- ELIMINAR PORTADA SI SE SOLICITA ---
            if request.POST.get('eliminar_portada') == '1':
                if formulario.imagen:
                    formulario.imagen.delete(save=False)
                    formulario.imagen = None
                    formulario.save()
            formulario.save()
            apariencia_form = FormularioAparienciaForm(instance=formulario)  # Refresca el form tras guardar
            response_data = {
                'success': True,
                'titulo': formulario.titulo,
                'subtitulo': formulario.subtitulo,
                'texto_boton_enviar': formulario.texto_boton_enviar
            }
            if formulario.imagen:
                response_data['imagen_url'] = formulario.imagen.url
            return JsonResponse(response_data)
        elif 'tipo' in request.POST:
            # Crear una nueva pregunta predefinida
            tipo = request.POST.get('tipo')
            texto = request.POST.get('texto', 'Nueva pregunta')
            descripcion = request.POST.get('descripcion', '')
            placeholder = request.POST.get('placeholder', '')
            pregunta = Pregunta.objects.create(
                formulario=formulario,
                tipo=tipo,
                texto=texto,
                descripcion=descripcion,
                placeholder=placeholder if tipo == 'texto' else '',
                pregunta_dependiente_id=request.POST.get('pregunta_dependiente') or None
            )
            
            # Inicializar valores por defecto según el tipo
            if tipo == 'opcion_multiple':
                pregunta.mostrar_como_radio = False
                pregunta.permitir_multiple = False
                pregunta.save()
                # Crear opciones por defecto
                Opcion.objects.create(pregunta=pregunta, texto="Opción 1")
                Opcion.objects.create(pregunta=pregunta, texto="Opción 2")
                Opcion.objects.create(pregunta=pregunta, texto="Opción 3")
            elif tipo == 'valoracion':
                pregunta.rango_minimo = 1
                pregunta.rango_maximo = 5
                pregunta.save()
            elif tipo == 'verdadero_falso':
                pregunta.mostrar_como_radio = True
                pregunta.save()
            
            # Si se está creando o editando una pregunta de tipo fecha, guardar el campo fecha_tipo
            if 'tipo' in request.POST and request.POST['tipo'] == 'fecha':
                pregunta.fecha_tipo = request.POST.get('fecha_tipo', 'envio')
                pregunta.save()
            
            return JsonResponse({
                'success': True,
                'pregunta_id': pregunta.id,
                'texto': pregunta.texto,
                'tipo': pregunta.tipo,
                'descripcion': pregunta.descripcion,
                'placeholder': pregunta.placeholder,
                'opciones': list(pregunta.opcion_set.values_list('texto', flat=True)) if pregunta.tipo == 'opcion_multiple' else [],
                'rango_minimo': pregunta.rango_minimo if pregunta.tipo == 'valoracion' else None,
                'rango_maximo': pregunta.rango_maximo if pregunta.tipo == 'valoracion' else None,
                'mostrar_como_radio': pregunta.mostrar_como_radio,
                'permitir_multiple': pregunta.permitir_multiple,
                'fecha_tipo': pregunta.fecha_tipo if pregunta.tipo == 'fecha' else None
            })
        else:
            # Actualizar una pregunta existente
            pregunta_id = request.POST.get('pregunta_id')
            if pregunta_id:
                pregunta = get_object_or_404(Pregunta, id=pregunta_id)
                pregunta.texto = request.POST.get('texto')
                pregunta.descripcion = request.POST.get('descripcion', '')
                pregunta.placeholder = request.POST.get('placeholder', '') if request.POST.get('tipo') == 'texto' else ''
                pregunta.tipo = request.POST.get('tipo')
                pregunta.mostrar_como_radio = request.POST.get('mostrar_como_radio') == 'on'
                pregunta.permitir_multiple = request.POST.get('permitir_multiple') == 'on' and pregunta.mostrar_como_radio
                pregunta.pregunta_dependiente_id = request.POST.get('pregunta_dependiente') or None
                
                if pregunta.tipo == 'valoracion':
                    pregunta.rango_minimo = request.POST.get('rango_minimo')
                    pregunta.rango_maximo = request.POST.get('rango_maximo')
                
                # Si se está creando o editando una pregunta de tipo fecha, guardar el campo fecha_tipo
                if 'tipo' in request.POST and request.POST['tipo'] == 'fecha':
                    pregunta.fecha_tipo = request.POST.get('fecha_tipo', 'envio')
                
                # NUEVO: Guardar mapa_tipo si es tipo mapa
                if pregunta.tipo == 'mapa':
                    pregunta.mapa_tipo = request.POST.get('mapa_tipo', 'punto')
                pregunta.save()
                
                if pregunta.tipo == 'opcion_multiple':
                    opciones = request.POST.get('opciones', '').split('\n')
                    pregunta.opcion_set.all().delete()
                    for opcion_texto in opciones:
                        if opcion_texto.strip():
                            Opcion.objects.create(pregunta=pregunta, texto=opcion_texto.strip())
                
                return JsonResponse({
                    'success': True,
                    'mostrar_como_radio': pregunta.mostrar_como_radio,
                    'permitir_multiple': pregunta.permitir_multiple,
                    'descripcion': pregunta.descripcion,
                    'placeholder': pregunta.placeholder,
                    'opciones': list(pregunta.opcion_set.values_list('texto', flat=True)) if pregunta.tipo == 'opcion_multiple' else [],
                    'rango_minimo': pregunta.rango_minimo if pregunta.tipo == 'valoracion' else None,
                    'rango_maximo': pregunta.rango_maximo if pregunta.tipo == 'valoracion' else None,
                    'fecha_tipo': pregunta.fecha_tipo if pregunta.tipo == 'fecha' else None
                })
    
    # Filtrar preguntas por el formulario actual
    preguntas = formulario.preguntas.all().order_by('id')
    for pregunta in preguntas:
        if pregunta.tipo == 'opcion_multiple':
            pregunta.opciones = Opcion.objects.filter(pregunta=pregunta)
        elif pregunta.tipo == 'valoracion':
            pregunta.rango = list(range(pregunta.rango_minimo, pregunta.rango_maximo + 1))
        # Asegurar que obligatorio sea booleano
        pregunta.obligatorio = bool(getattr(pregunta, 'obligatorio', False))
        if pregunta.tipo == 'fecha':
            # Guardar el tipo de fecha (envio o usuario) en el modelo si lo tienes, o como atributo extra
            pregunta.fecha_tipo = getattr(pregunta, 'fecha_tipo', 'envio')

    return render(request, 'agregar_pregunta.html', {
        'formulario': formulario,
        'preguntas': preguntas,
        'apariencia_form': apariencia_form,
        'color_temas': color_temas,
        'departamentos': departamentos,
    })

def formulario_con_mapa(request):
    puntos = Punto.objects.all()
    puntos_json = serializers.serialize('json', puntos)
    preguntas = Pregunta.objects.all()

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        direccion = request.POST.get('direccion')
        lat = request.POST.get('lat')
        lng = request.POST.get('lng')
        punto = Punto.objects.create(nombre=nombre, direccion=direccion, latitud=lat, longitud=lng)

        for pregunta in preguntas:
            respuesta_texto = request.POST.get(f'respuesta_{pregunta.id}')
            if respuesta_texto:
                Respuesta.objects.create(pregunta=pregunta, punto=punto, respuesta=respuesta_texto)
        return redirect('ver_guardado')

    return render(request, 'guardar_punto.html', {'puntos_json': puntos_json, 'preguntas': preguntas})

def cargar_imagen(request):
    if request.method == 'POST':
        formulario_id = request.POST.get('formulario_id')
        formulario = get_object_or_404(Formulario, id=formulario_id)
        
        if 'imagen' in request.FILES:
            # Eliminar la imagen anterior si existe
            if formulario.imagen:
                formulario.imagen.delete(save=True)
                formulario.imagen = None
                formulario.save()
            # Guardar la nueva imagen
            formulario.imagen = request.FILES['imagen']
            formulario.save()
            messages.success(request, 'Imagen cargada correctamente.')
        else:
            messages.error(request, 'No se ha seleccionado ninguna imagen.')
        
        return redirect('agregar_pregunta', formulario_id=formulario_id)
    return redirect('agregar_pregunta', formulario_id=request.POST.get('formulario_id'))

def eliminar_imagen_formulario(request, formulario_id):
    if request.method == 'POST':
        formulario = get_object_or_404(Formulario, id=formulario_id)
        if formulario.imagen:
            formulario.imagen.delete(save=True)
            formulario.imagen = None
            formulario.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)

def actualizar_preguntas(request):
    formulario_hashid = request.GET.get('formulario_id')
    formulario_id = decode_hashid_or_404(formulario_hashid)
    preguntas = Pregunta.objects.filter(formulario_id=formulario_id).order_by('id')
    # Preparar las preguntas igual que en guardar_punto
    for pregunta in preguntas:
        if pregunta.tipo == 'opcion_multiple':
            pregunta.opciones = Opcion.objects.filter(pregunta=pregunta)
        elif pregunta.tipo == 'valoracion':
            pregunta.rango = list(range(pregunta.rango_minimo, pregunta.rango_maximo + 1))
        pregunta.obligatorio = bool(getattr(pregunta, 'obligatorio', False))
        if pregunta.tipo == 'fecha':
            pregunta.fecha_tipo = getattr(pregunta, 'fecha_tipo', 'envio')
    return render(request, 'questions_fragment.html', {
        'preguntas': preguntas,
        'mostrar_boton_eliminar': True
    })

def eliminar_pregunta(request, pregunta_id):
    if request.method == 'DELETE':
        try:
            pregunta = Pregunta.objects.get(id=pregunta_id)
            pregunta.delete()
            return JsonResponse({'success': True, 'message': 'Pregunta eliminada correctamente.'})
        except Pregunta.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'La pregunta no existe.'}, status=404)
    return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)

def obtener_pregunta(request, pregunta_id):
    try:
        pregunta = Pregunta.objects.get(id=pregunta_id)
        # Refuerzo: asegurar que mapa_tipo siempre tenga un valor válido si es tipo mapa
        mapa_tipo = None
        if pregunta.tipo == 'mapa':
            mapa_tipo = getattr(pregunta, 'mapa_tipo', None)
            if mapa_tipo not in ['punto', 'linea', 'poligono']:
                mapa_tipo = 'punto'
            # DEBUG: log para ver el valor recibido
            import logging
            logger = logging.getLogger('django')
            logger.info(f"[DEBUG obtener_pregunta] pregunta.id={pregunta.id} mapa_tipo={mapa_tipo}")
        data = {
            'texto': pregunta.texto,
            'tipo': pregunta.tipo,
            'descripcion': pregunta.descripcion,
            'placeholder': pregunta.placeholder,
            'opciones': list(pregunta.opcion_set.values_list('texto', flat=True)) if pregunta.tipo == 'opcion_multiple' else [],
            'rango_minimo': pregunta.rango_minimo if pregunta.tipo == 'valoracion' else None,
            'rango_maximo': pregunta.rango_maximo if pregunta.tipo == 'valoracion' else None,
            'mostrar_como_radio': pregunta.mostrar_como_radio,
            'permitir_multiple': pregunta.permitir_multiple,
            'obligatorio': pregunta.obligatorio,
            'fecha_tipo': pregunta.fecha_tipo if pregunta.tipo == 'fecha' else None,
            'mapa_tipo': mapa_tipo if pregunta.tipo == 'mapa' else None,
            'pregunta_dependiente': pregunta.pregunta_dependiente_id
        }
        return JsonResponse(data)
    except Pregunta.DoesNotExist:
        return JsonResponse({'error': 'Pregunta no encontrada'}, status=404)
    
def editar_pregunta(request, pregunta_id):
    if request.method == 'POST':
        try:
            pregunta = Pregunta.objects.get(id=pregunta_id)
            pregunta.texto = request.POST.get('texto')
            pregunta.descripcion = request.POST.get('descripcion', '')
            pregunta.placeholder = request.POST.get('placeholder', '') if request.POST.get('tipo') == 'texto' else ''
            pregunta.tipo = request.POST.get('tipo')
            pregunta.mostrar_como_radio = request.POST.get('mostrar_como_radio') == 'on'
            pregunta.permitir_multiple = request.POST.get('permitir_multiple') == 'on' and pregunta.mostrar_como_radio
            pregunta.pregunta_dependiente_id = request.POST.get('pregunta_dependiente') or None
            
            if pregunta.tipo == 'valoracion':
                pregunta.rango_minimo = request.POST.get('rango_minimo')
                pregunta.rango_maximo = request.POST.get('rango_maximo')
            
            if 'tipo' in request.POST and request.POST['tipo'] == 'fecha':
                pregunta.fecha_tipo = request.POST.get('fecha_tipo', 'envio')
            
            # NUEVO: Guardar mapa_tipo si es tipo mapa
            if pregunta.tipo == 'mapa':
                pregunta.mapa_tipo = request.POST.get('mapa_tipo', 'punto')
            
            pregunta.save()
            
            if pregunta.tipo == 'opcion_multiple':
                # Actualizar opciones
                opciones = request.POST.get('opciones', '').split('\n')
                pregunta.opcion_set.all().delete()
                for opcion_texto in opciones:
                    if opcion_texto.strip():
                        Opcion.objects.create(pregunta=pregunta, texto=opcion_texto.strip())
            
            return JsonResponse({
                'success': True,
                'mostrar_como_radio': pregunta.mostrar_como_radio,
                'permitir_multiple': pregunta.permitir_multiple,
                'descripcion': pregunta.descripcion,
                'placeholder': pregunta.placeholder,
                'opciones': list(pregunta.opcion_set.values_list('texto', flat=True)) if pregunta.tipo == 'opcion_multiple' else [],
                'rango_minimo': pregunta.rango_minimo if pregunta.tipo == 'valoracion' else None,
                'rango_maximo': pregunta.rango_maximo if pregunta.tipo == 'valoracion' else None,
                'fecha_tipo': pregunta.fecha_tipo if pregunta.tipo == 'fecha' else None
            })
        except Pregunta.DoesNotExist:
            return JsonResponse({'error': 'Pregunta no encontrada'}, status=404)
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required(login_url='/login/')
def administrar_formularios(request):
    if request.method == 'POST':
        form = FormularioForm(request.POST)
        if form.is_valid():
            try:
                formulario = form.save()
                # Asegurarnos de que el formulario se guardó correctamente
                if formulario.id:
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': True, 
                            'redirect_url': reverse('agregar_pregunta', kwargs={'formulario_hashid': hashids.encode(formulario.id)})
                        })
                    return redirect('agregar_pregunta', formulario_hashid=hashids.encode(formulario.id))
                else:
                    messages.error(request, 'Error al guardar el formulario')
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': False,
                            'error': 'Error al guardar el formulario'
                        }, status=400)
            except IntegrityError:
                messages.error(request, 'El nombre del formulario ya existe. Por favor, elige otro nombre.')
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'error': 'El nombre del formulario ya existe. Por favor, elige otro nombre.'
                    }, status=400)
    else:
        form = FormularioForm()

    formularios = Formulario.objects.all()
    return render(request, 'administrar_formularios.html', {
        'form': form,
        'formularios': formularios,
        'show_errors': request.method == 'POST'
    })

def eliminar_formulario(request, formulario_id):
    formulario = get_object_or_404(Formulario, id=formulario_id)
    formulario.delete()
    return JsonResponse({'success': True, 'message': 'Formulario eliminado correctamente.'})

# Añadir esta nueva vista al archivo views.py
def eliminar_punto(request, punto_id):
    if request.method == 'DELETE':
        try:
            punto = Punto.objects.get(id=punto_id)
            
            # Eliminar las respuestas asociadas a este punto
            Respuesta.objects.filter(punto=punto).delete()
            
            # Eliminar las imágenes asociadas a este punto
            ImagenRespuesta.objects.filter(punto=punto).delete()
            
            # Finalmente eliminar el punto
            punto.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Punto eliminado correctamente'
            })
        except Punto.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Punto no encontrado'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Método no permitido'
    }, status=405)

def guardar_cambios_formulario(request):
    if request.method == 'POST':
        try:
            # Soportar JSON puro en el body
            if request.content_type == 'application/json':
                data = json.loads(request.body)
                cambios = data.get('cambios', {})
                formulario_hashid = cambios.get('formulario_id')
                estilos_header = cambios.get('estilos_header', {})
                orden_preguntas = cambios.get('orden_preguntas', [])
            else:
                cambios = json.loads(request.POST.get('cambios', '{}'))
                formulario_hashid = request.POST.get('formulario_id')
                estilos_header = json.loads(request.POST.get('estilos_header', '{}'))
                orden_preguntas = json.loads(request.POST.get('orden_preguntas', '[]'))
            formulario_id = decode_hashid_or_404(formulario_hashid)
            formulario = get_object_or_404(Formulario, id=formulario_id)
            # === ELIMINAR PORTADA SI SE SOLICITA ===
            eliminar_portada = request.POST.get('eliminar_portada')
            if eliminar_portada == '1':
                if formulario.imagen:
                    formulario.imagen.delete(save=False)
                    formulario.imagen = None
                    formulario.save()
            # === GUARDAR TÍTULO Y SUBTÍTULO DEL FORMULARIO SI SE ENVÍAN ===
            titulo = request.POST.get('titulo')
            subtitulo = request.POST.get('subtitulo')
            # Si el título es None o vacío, poner 'Título' por defecto
            if not titulo:
                titulo = 'Título'
            formulario.titulo = titulo
            if subtitulo is not None:
                formulario.subtitulo = subtitulo
            formulario.save()

            # === GUARDAR IMAGEN DE PORTADA SI SE ENVÍA ===
            # LOG para depuración de imagen
            print('DEBUG request.FILES:', request.FILES)
            if 'imagen' in request.FILES:
                print('DEBUG imagen recibida:', request.FILES['imagen'])
                if formulario.imagen:
                    print('DEBUG eliminando imagen anterior:', formulario.imagen)
                    formulario.imagen.delete()
                formulario.imagen = request.FILES['imagen']
                formulario.save()
                print('DEBUG imagen guardada en formulario:', formulario.imagen)
            else:
                print('DEBUG no se recibió imagen en el POST')
            # === GUARDAR TEXTO DEL BOTÓN ENVIAR SI SE ENVÍA ===
            texto_boton_enviar = request.POST.get('texto_boton_enviar')
            if texto_boton_enviar is not None:
                formulario.texto_boton_enviar = texto_boton_enviar
                formulario.save()
            
            # === GUARDAR COLOR DE TEMA SI SE ENVÍA ===
            tema_color = request.POST.get('tema_color')
            if tema_color is not None:
                formulario.tema_color = tema_color
            formulario.save()
            
            # Actualizar estilos del formulario
            if estilos_header:
                titulo_estilos = estilos_header.get('titulo', {})
                subtitulo_estilos = estilos_header.get('subtitulo', {})
                # Actualizar estilos del título
                formulario.titulo_bold = titulo_estilos.get('bold', False)
                formulario.titulo_italic = titulo_estilos.get('italic', False)
                formulario.titulo_underline = titulo_estilos.get('underline', False)
                formulario.titulo_color = titulo_estilos.get('color', '#000000')
                formulario.titulo_font = titulo_estilos.get('font', "'Inter',sans-serif")
                formulario.titulo_align = titulo_estilos.get('align', 'center')
                # Actualizar estilos del subtítulo
                formulario.subtitulo_bold = subtitulo_estilos.get('bold', False)
                formulario.subtitulo_italic = subtitulo_estilos.get('italic', False)
                formulario.subtitulo_underline = subtitulo_estilos.get('underline', False)
                formulario.subtitulo_color = subtitulo_estilos.get('color', '#000000')
                formulario.subtitulo_font = subtitulo_estilos.get('font', "'Inter',sans-serif")
                formulario.subtitulo_align = subtitulo_estilos.get('align', 'center')
                formulario.subtitulo_fontSize = subtitulo_estilos.get('fontSize', '16px') # <--- ¡AÑADE ESTA LÍNEA!
                formulario.save()
            
            # Procesar eliminaciones
            for pregunta_id in cambios.get('eliminadas', []):
                try:
                    pregunta = Pregunta.objects.get(id=pregunta_id)
                    pregunta.delete()
                except Pregunta.DoesNotExist:
                    pass
            # Procesar ediciones
            for pregunta_data in cambios.get('editadas', []):
                try:
                    pregunta = Pregunta.objects.get(id=pregunta_data['id'])
                    pregunta.texto = pregunta_data.get('texto', pregunta.texto)
                    pregunta.descripcion = pregunta_data.get('descripcion', pregunta.descripcion)
                    pregunta.placeholder = pregunta_data.get('placeholder', pregunta.placeholder) if pregunta_data.get('tipo') == 'texto' else ''
                    pregunta.tipo = pregunta_data.get('tipo', pregunta.tipo)
                    pregunta.mostrar_como_radio = pregunta_data.get('mostrar_como_radio', pregunta.mostrar_como_radio)
                    pregunta.permitir_multiple = pregunta_data.get('permitir_multiple', pregunta.permitir_multiple)
                    pregunta.obligatorio = pregunta_data.get('obligatorio', False)
                    pregunta.pregunta_dependiente_id = pregunta_data.get('pregunta_dependiente') or None
                    if pregunta.tipo == 'valoracion':
                        pregunta.rango_minimo = pregunta_data.get('rango_minimo', pregunta.rango_minimo)
                        pregunta.rango_maximo = pregunta_data.get('rango_maximo', pregunta.rango_maximo)
                    if pregunta.tipo == 'fecha':
                        pregunta.fecha_tipo = pregunta_data.get('fecha_tipo', 'envio')
                    # NUEVO: Guardar mapa_tipo si es tipo mapa
                    if pregunta.tipo == 'mapa':
                        pregunta.mapa_tipo = pregunta_data.get('mapa_tipo', 'punto')
                    pregunta.save()
                    if pregunta.tipo == 'opcion_multiple':
                        opciones = pregunta_data.get('opciones', [])
                        pregunta.opcion_set.all().delete()
                        for opcion_texto in opciones:
                            if opcion_texto.strip():
                                Opcion.objects.create(pregunta=pregunta, texto=opcion_texto.strip())
                except Pregunta.DoesNotExist:
                    pass
            # Procesar creaciones
            for pregunta_data in cambios.get('creadas', []):
                # Obtener el mayor orden actual para el formulario
                ultimo_orden = (
                    Pregunta.objects.filter(formulario=formulario)
                    .aggregate(max_orden=models.Max('orden'))
                    .get('max_orden')
                )
                if ultimo_orden is None:
                    ultimo_orden = -1
                nuevo_orden = ultimo_orden + 1

                pregunta = Pregunta.objects.create(
                    formulario=formulario,
                    tipo=pregunta_data['tipo'],
                    texto=pregunta_data['texto'],
                    descripcion=pregunta_data.get('descripcion', ''),
                    placeholder=pregunta_data.get('placeholder', '') if pregunta_data.get('tipo') == 'texto' else '',
                    obligatorio=pregunta_data.get('obligatorio', False),
                    orden=nuevo_orden,
                    mapa_tipo=pregunta_data.get('mapa_tipo', 'punto') if pregunta_data.get('tipo') == 'mapa' else 'punto',
                    pregunta_dependiente_id=pregunta_data.get('pregunta_dependiente') or None
                )
                if pregunta.tipo == 'opcion_multiple':
                    pregunta.mostrar_como_radio = pregunta_data.get('mostrar_como_radio', False)
                    pregunta.permitir_multiple = pregunta_data.get('permitir_multiple', False)
                    pregunta.save()
                    for opcion_texto in pregunta_data.get('opciones', []):
                        if opcion_texto.strip():
                            Opcion.objects.create(pregunta=pregunta, texto=opcion_texto.strip())
                elif pregunta.tipo == 'valoracion':
                    pregunta.rango_minimo = pregunta_data.get('rango_minimo', 1)
                    pregunta.rango_maximo = pregunta_data.get('rango_maximo', 5)
                    pregunta.save()
                elif pregunta.tipo == 'fecha':
                    pregunta.fecha_tipo = pregunta_data.get('fecha_tipo', 'envio')
                    pregunta.save()

            # === ACTUALIZAR ORDEN DE PREGUNTAS SI SE ENVÍA ===
            orden_preguntas = None
            if request.content_type == 'application/json':
                orden_preguntas = data.get('orden_preguntas')
            else:
                orden_preguntas = request.POST.getlist('orden_preguntas[]') or json.loads(request.POST.get('orden_preguntas', '[]'))
            if orden_preguntas:
                for idx, pregunta_id in enumerate(orden_preguntas):
                    try:
                        pregunta = Pregunta.objects.get(id=pregunta_id)
                        pregunta.orden = idx
                        pregunta.save(update_fields=['orden'])
                    except Pregunta.DoesNotExist:
                        continue

            permitir_volver_a_contestar = request.POST.get('permitir_volver_a_contestar')
            if permitir_volver_a_contestar is not None:
                formulario.permitir_volver_a_contestar = permitir_volver_a_contestar in ['1', 'true', 'True', True]
                formulario.save()

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

def descargar_excel(request, formulario_hashid):
    formulario_id = decode_hashid_or_404(formulario_hashid)
    formulario = get_object_or_404(Formulario, id=formulario_id)
    preguntas = formulario.preguntas.all()
    puntos = Punto.objects.filter(respuesta__pregunta__formulario=formulario).distinct().order_by('id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"

    # Solo columnas: ID, Latitud, Longitud, y preguntas (excluyendo tipo 'mapa')
    preguntas_sin_mapa = [p for p in preguntas if p.tipo != 'mapa']
    headers = ['ID', 'Latitud', 'Longitud'] + [p.texto for p in preguntas_sin_mapa]
    ws.append(headers)

    if not puntos.exists():
        ws.append(["No hay puntos guardados"])
    else:
        for punto in puntos:
            fila = [
                punto.id,
                punto.latitud,
                punto.longitud
            ]
            for pregunta in preguntas_sin_mapa:
                respuestas = Respuesta.objects.filter(punto=punto, pregunta=pregunta)
                if pregunta.tipo == 'foto' and respuestas.exists():
                    imagenes = []
                    for r in respuestas:
                        if r.imagen_respuesta and hasattr(r.imagen_respuesta.imagen, 'url'):
                            url = request.build_absolute_uri(r.imagen_respuesta.imagen.url)
                            imagenes.append(url)
                        else:
                            imagenes.append(r.respuesta or '')
                    # Si solo hay una imagen, poner el string directamente
                    if len(imagenes) == 1:
                        fila.append(imagenes[0])
                    elif len(imagenes) > 1:
                        fila.append(', '.join(imagenes))
                    else:
                        fila.append('')
                else:
                    respuestas_texto = [r.respuesta for r in respuestas] if respuestas.exists() else []
                    # Si solo hay una respuesta, poner el string directamente
                    if len(respuestas_texto) == 1:
                        fila.append(respuestas_texto[0])
                    elif len(respuestas_texto) > 1:
                        fila.append(', '.join(respuestas_texto))
                    else:
                        fila.append('')
            ws.append(fila)

        # Convertir las celdas de imágenes a hipervínculos
        from urllib.parse import urlparse
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row[3:]:
                if isinstance(cell.value, str) and cell.value.startswith('http'):
                    cell.hyperlink = cell.value
                    parsed = urlparse(cell.value)
                    filename = parsed.path.split('/')[-1] if '/' in parsed.path else cell.value
                    cell.value = filename

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=datos_formulario_{formulario_hashid}.xlsx'
    return response


def descargar_excel(request, formulario_hashid):
    formulario_id = decode_hashid_or_404(formulario_hashid)
    formulario = get_object_or_404(Formulario, id=formulario_id)
    preguntas = formulario.preguntas.all()
    puntos = Punto.objects.filter(respuesta__pregunta__formulario=formulario).distinct().order_by('id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"

    # Solo columnas: ID, Latitud, Longitud, y preguntas (excluyendo tipo 'mapa')
    preguntas_sin_mapa = [p for p in preguntas if p.tipo != 'mapa']
    headers = ['ID', 'Latitud', 'Longitud'] + [p.texto for p in preguntas_sin_mapa]
    ws.append(headers)

    if not puntos.exists():
        ws.append(["No hay puntos guardados"])
    else:
        for punto in puntos:
            fila = [
                punto.id,
                punto.latitud,
                punto.longitud
            ]
            for pregunta in preguntas_sin_mapa:
                respuestas = Respuesta.objects.filter(punto=punto, pregunta=pregunta)
                if pregunta.tipo == 'foto' and respuestas.exists():
                    imagenes = []
                    for r in respuestas:
                        if r.imagen_respuesta and hasattr(r.imagen_respuesta.imagen, 'url'):
                            url = request.build_absolute_uri(r.imagen_respuesta.imagen.url)
                            imagenes.append(url)
                        else:
                            imagenes.append(r.respuesta or '')
                    # Si solo hay una imagen, poner el string directamente
                    if len(imagenes) == 1:
                        fila.append(imagenes[0])
                    elif len(imagenes) > 1:
                        fila.append(', '.join(imagenes))
                    else:
                        fila.append('')
                else:
                    respuestas_texto = [r.respuesta for r in respuestas] if respuestas.exists() else []
                    # Si solo hay una respuesta, poner el string directamente
                    if len(respuestas_texto) == 1:
                        fila.append(respuestas_texto[0])
                    elif len(respuestas_texto) > 1:
                        fila.append(', '.join(respuestas_texto))
                    else:
                        fila.append('')
            ws.append(fila)

        # Convertir las celdas de imágenes a hipervínculos
        from urllib.parse import urlparse
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row[3:]:
                if isinstance(cell.value, str) and cell.value.startswith('http'):
                    cell.hyperlink = cell.value
                    parsed = urlparse(cell.value)
                    filename = parsed.path.split('/')[-1] if '/' in parsed.path else cell.value
                    cell.value = filename

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=datos_formulario_{formulario_hashid}.xlsx'
    return response

def descargar_kmz(request, formulario_hashid):
    formulario_id = decode_hashid_or_404(formulario_hashid)
    formulario = get_object_or_404(Formulario, id=formulario_id)
    preguntas = formulario.preguntas.all()
    puntos = Punto.objects.filter(respuesta__pregunta__formulario=formulario).distinct().order_by('id')

    # Crear KML
    kml = Element('kml', xmlns="http://www.opengis.net/kml/2.2")
    doc = SubElement(kml, 'Document')
    if not puntos.exists():
        SubElement(doc, 'Placemark').text = 'No hay puntos guardados.'
    else:
        for punto in puntos:
            placemark = SubElement(doc, 'Placemark')
            SubElement(placemark, 'name').text = getattr(punto, 'nombre', f"Punto {punto.id}")
            desc = ""
            for pregunta in preguntas:
                respuestas = Respuesta.objects.filter(punto=punto, pregunta=pregunta)
                if respuestas.exists():
                    if pregunta.tipo == 'foto':
                        # Mostrar la imagen como <img src="URL"> si existe
                        for r in respuestas:
                            if r.imagen_respuesta and hasattr(r.imagen_respuesta.imagen, 'url'):
                                url = request.build_absolute_uri(r.imagen_respuesta.imagen.url)
                                desc += f"<b>{pregunta.texto}:</b> <br><img src='{url}' style='max-width:300px;'><br/>"
                            else:
                                desc += f"<b>{pregunta.texto}:</b> {r.respuesta}<br/>"
                    else:
                        desc += f"<b>{pregunta.texto}:</b> {', '.join([r.respuesta for r in respuestas])}<br/>"
            SubElement(placemark, 'description').text = desc
            point = SubElement(placemark, 'Point')
            SubElement(point, 'coordinates').text = f"{punto.longitud},{punto.latitud},0"

    kml_bytes = tostring(kml, encoding='utf-8', method='xml')
    kmz_buffer = BytesIO()
    with zipfile.ZipFile(kmz_buffer, 'w', zipfile.ZIP_DEFLATED) as kmz:
        kmz.writestr('doc.kml', kml_bytes)
    kmz_buffer.seek(0)
    response = HttpResponse(kmz_buffer.getvalue(), content_type='application/vnd.google-earth.kmz')
    response['Content-Disposition'] = f'attachment; filename=datos_formulario_{formulario_hashid}.kmz'
    return response

def obtener_datos_formulario(request, formulario_hashid):
    formulario_id = decode_hashid_or_404(formulario_hashid)
    formulario = get_object_or_404(Formulario, id=formulario_id)

    # --- NUEVO: Forzar valores por defecto en el backend para API ---
    titulo = formulario.titulo
    subtitulo = formulario.subtitulo
    if not titulo or titulo == 'none':
        titulo = 'titulo'
    if not subtitulo or subtitulo == 'none':
        subtitulo = 'subtitulo'

    # Obtener los estilos del header
    estilos_header = {
        'titulo': {
            'bold': formulario.titulo_bold,
            'italic': formulario.titulo_italic,
            'underline': formulario.titulo_underline,
            'color': formulario.titulo_color,
            'font': formulario.titulo_font,
            'align': formulario.titulo_align
        },
        'subtitulo': {
            'bold': formulario.subtitulo_bold,
            'italic': formulario.subtitulo_italic,
            'underline': formulario.subtitulo_underline,
            'color': formulario.subtitulo_color,
            'font': formulario.subtitulo_font,
            'align': formulario.subtitulo_align,
            'fontSize': formulario.subtitulo_fontSize # <--- ¡AÑADE ESTA LÍNEA!
        }
    }
    
    preguntas = formulario.preguntas.all().order_by('id')
    preguntas_data = []
    
    for pregunta in preguntas:
        pregunta_data = {
            'id': pregunta.id,
            'texto': pregunta.texto,
            'tipo': pregunta.tipo,
            'descripcion': pregunta.descripcion,
            'placeholder': pregunta.placeholder,
            'mostrar_como_radio': pregunta.mostrar_como_radio,
            'permitir_multiple': pregunta.permitir_multiple
        }
        
        if pregunta.tipo == 'opcion_multiple':
            pregunta_data['opciones'] = list(pregunta.opcion_set.values_list('texto', flat=True))
        elif pregunta.tipo == 'valoracion':
            pregunta_data['rango_minimo'] = pregunta.rango_minimo
            pregunta_data['rango_maximo'] = pregunta.rango_maximo
        if pregunta.tipo == 'fecha':
            pregunta_data['fecha_tipo'] = pregunta.fecha_tipo
            
        preguntas_data.append(pregunta_data)
    
    return JsonResponse({
        'id': formulario.id,
        'titulo': titulo,
        'subtitulo': subtitulo,
        'estilos_header': estilos_header,
        'preguntas': preguntas_data
    })

@csrf_exempt
def actualizar_portada_formulario(request, formulario_id):
    print("ENTRANDO A LA VISTA actualizar_portada_formulario")  # DEBUG
    logger = logging.getLogger("django")
    logger.info(f"[actualizar_portada_formulario] Método: {request.method}")
    try:
        if request.method == 'POST':
            formulario = get_object_or_404(Formulario, id=formulario_id)
            logger.info(f"[actualizar_portada_formulario] Archivos recibidos: {request.FILES}")
            print(f"Archivos recibidos: {request.FILES}")  # DEBUG
            if 'imagen' in request.FILES:
                try:
                    if formulario.imagen:
                        logger.info("[actualizar_portada_formulario] Eliminando imagen anterior")
                        print("Eliminando imagen anterior")  # DEBUG
                        formulario.imagen.delete()
                    formulario.imagen = request.FILES['imagen']
                    formulario.save()
                    logger.info(f"[actualizar_portada_formulario] Imagen guardada correctamente: {getattr(formulario.imagen, 'url', 'NO_URL')}")
                    print(f"Imagen guardada correctamente: {getattr(formulario.imagen, 'url', 'NO_URL')}")  # DEBUG
                    if hasattr(formulario.imagen, 'url'):
                        return JsonResponse({'success': True, 'imagen_url': formulario.imagen.url})
                    else:
                        return JsonResponse({'success': True, 'imagen_url': ''})
                except Exception as e:
                    logger.error(f"[actualizar_portada_formulario] Error al guardar la imagen: {e}")
                    print(f"Error al guardar la imagen: {e}")  # DEBUG
                    return JsonResponse({'success': False, 'error': f'Error al guardar la imagen: {e}'})
            logger.warning("[actualizar_portada_formulario] No se recibió imagen en la petición")
            print("No se recibió imagen en la petición")  # DEBUG
            return JsonResponse({'success': False, 'error': 'No se recibió imagen'})
        logger.warning(f"[actualizar_portada_formulario] Método no permitido: {request.method}")
        print(f"Método no permitido: {request.method}")  # DEBUG
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    except Exception as e:
        logger.error(f"[actualizar_portada_formulario] Excepción inesperada: {e}")
        print(f"Excepción inesperada: {e}")  
        return JsonResponse({'success': False, 'error': f'Excepción inesperada: {e}'})

@csrf_exempt
def actualizar_orden_preguntas(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            orden = data.get('orden', [])
            for idx, pregunta_id in enumerate(orden):
                try:
                    pregunta = Pregunta.objects.get(id=pregunta_id)
                    pregunta.orden = idx
                    pregunta.save(update_fields=['orden'])
                except Pregunta.DoesNotExist:
                    continue
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})



@csrf_exempt
@require_POST
def actualizar_nombre_formulario(request, formulario_hashid):
    try:
        from .views import decode_hashid_or_404
        formulario_id = decode_hashid_or_404(formulario_hashid)
        data = json.loads(request.body)
        nombre = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '').strip()
        departamento_id = data.get('departamento', '').strip()
        if not nombre:
            return JsonResponse({'success': False, 'error': 'Nombre requerido'})
        formulario = Formulario.objects.get(id=formulario_id)
        formulario.nombre = nombre
        formulario.descripcion = descripcion
        if hasattr(formulario, 'departamento'):
            if departamento_id:
                try:
                    departamento = Departamento.objects.get(id=departamento_id)
                except Departamento.DoesNotExist:
                    departamento = None
                formulario.departamento = departamento
            else:
                formulario.departamento = None
            formulario.save(update_fields=['nombre', 'descripcion', 'departamento'])
        else:
            formulario.save(update_fields=['nombre', 'descripcion'])
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
@require_POST
def actualizar_punto(request, punto_id):
    try:
        punto = Punto.objects.get(id=punto_id)
        data = json.loads(request.body)
        lat = float(data.get('latitud'))
        lng = float(data.get('longitud'))
        punto.latitud = lat
        punto.longitud = lng
        punto.save()
        return JsonResponse({'success': True})
    except Punto.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Punto no encontrado.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)
