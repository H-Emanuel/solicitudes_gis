import json
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,update_session_auth_hash, login as auth_login, logout as auth_logout
from django.shortcuts import redirect
from django.contrib.auth.models import User
from django.http import  HttpResponse,JsonResponse
from core.models import UserActivity
from formulario.models import ProtocoloSolicitud,Registro_designio
from openpyxl import Workbook
# Create your views here.
from django.views.decorators.csrf import csrf_exempt
from .forms import ImagenForm,PDFForm
from .models import Imagen_sig,PDF_sig
from django.shortcuts import get_object_or_404, redirect
from django.core.paginator import Paginator
from datetime import timedelta
from formulario.models import *
from django.db.models import Q, Sum,Count
import os
from django.contrib import messages
from django.utils.timezone import now
from smtplib import *
from .pdf_generator import *
from tareas.models import *
from archivos.models import acceso_pagina


ESTADO = [
    ('RECIBIDO', 'RECIBIDO'),
    ('EN PROCESO', 'EN PROCESO'),
    ('EJECUTADO', 'EJECUTADO'),
    ('RECHAZADO', 'RECHAZADO')
]

LIMITE_DE_DIA = [
    ('', ''),
    ('L', '1 - 2 Días Hábiles'),
    ('M', '2 - 4 Días Hábiles'),
    ('A', '3 - 5 Días Hábiles'),
    ('P', 'Especificar días manualmente'),
]

OPCIONES = {
    'ESTADO': ESTADO,
    'LIMITE_DE_DIA': LIMITE_DE_DIA
}


@csrf_exempt
def login(request):
    next_url = request.GET.get('next') or request.POST.get('next') or 'control'
    if request.user.is_authenticated:
        # Si el usuario tiene acceso_pagina.acceso=True, redirigir a otra página
        if acceso_pagina.objects.filter(usuario=request.user, acceso=True).exists():
            return redirect('pagina_sin_acceso')  # Cambia 'pagina_sin_acceso' por el nombre de tu URL
        return redirect(next_url)

    if request.method == 'POST':
        email = request.POST.get('email', '').lower()
        password = request.POST.get('password')
        user = authenticate(request, username=email, password=password)

        if user is not None:
            # Si el usuario tiene acceso_pagina.acceso=True, redirigir a otra página
            if acceso_pagina.objects.filter(usuario=user, acceso=True).exists():
                return redirect('inicio')  # Cambia 'pagina_sin_acceso' por el nombre de tu URL
            auth_login(request, user)
            next_url = request.GET.get('next') or request.POST.get('next') or 'control'
            return redirect(next_url)
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')

    # Pasa el parámetro 'next' al template para mantenerlo en el formulario
    return render(request, 'Login.html', {'next': request.GET.get('next', '')})

def download_excel(request):
    # Crear un nuevo libro de trabajo de Excel y agregar datos
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'PAGINA'
    sheet['B1'] = 'DEPARTAMENTO'
    sheet['C1'] = 'MES-DIA-AÑO-HORA'

    sheet.column_dimensions['A'].width = 30  # Ancho de la columna B
    sheet.column_dimensions['B'].width = 50  # Ancho de la columna C
    sheet.column_dimensions['C'].width = 30  # Ancho de la columna C


    row = 2  # Fila inicial para los datos
    activities = UserActivity.objects.all()

    for activity in activities:
        sheet.cell(row=row, column=1).value = activity.page
        sheet.cell(row=row, column=2).value = activity.departamento
        sheet.cell(row=row, column=3).value = activity.timestamp.strftime('%m-%d-%Y-%H:%M')
        row += 1  # Incrementar la fila para el próximo registro
    # Configurar la respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Historial_de_visitas.xlsx'

    # Guardar el libro de trabajo en la respuesta HTTP
    workbook.save(response)

    return response

def Historial_Visitas(request):
    Historial = UserActivity.objects.all()
    data = {
    'historial': Historial,
    }
    return render(request,'Historial_Visitas.html',data)

@login_required(login_url='/login/')
def solicitude_llegadas(request, dia_p=None):
    minutos = 0
    hora = 0
    usuarios = User.objects.exclude(acceso_pagina__acceso=True)

    # Filtrar las solicitudes según el usuario
    if request.user.is_superuser:
        solicitudes = ProtocoloSolicitud.objects.all()
    else:
        solicitudes = ProtocoloSolicitud.objects.filter(profesional=request.user)

    # Agregar información adicional sobre número de designios y días restantes
    solicitudes_data = []
    for solicitud in solicitudes:
        # Obtener el número de designios asociados a la solicitud
        numero_designios = Registro_designio.objects.filter(protocolo=solicitud).count()

        # Calcular los días restantes hasta la fecha límite
        if solicitud.fecha_T:
            dias_restantes = "Trabajo terminado"

            # solicitud.fecha_L el la fecha limite de la entrega 
        elif solicitud.fecha_L:
            # Calcular la diferencia en segundos
            total_segundos = (solicitud.fecha_L - now()).total_seconds()

            if total_segundos < 0:  # Si el tiempo ya pasó
                total_segundos = abs(total_segundos)
                dias_pasados = int(total_segundos // (24 * 3600))
                horas_pasadas = int((total_segundos % (24 * 3600)) // 3600)
                minutos_pasados = int((total_segundos % 3600) // 60)

                if dias_pasados > 0 or horas_pasadas > 0 or minutos_pasados > 0:
                    dias_restantes = f"Pasada por {dias_pasados} días, {horas_pasadas} horas y {minutos_pasados} minutos"
                else:
                    dias_restantes = "El tiempo límite ha pasado recientemente"
            else:  # Tiempo restante
                # Calcular días hábiles restantes
                dias_habiles_restantes = calcular_dias_habiles(now().date(), solicitud.fecha_L.date())
                horas_restantes = int((total_segundos % (24 * 3600)) // 3600)
                minutos_restantes = int((total_segundos % 3600) // 60)
                

                if dias_habiles_restantes > 1:
                    dias_restantes = f"Te quedan {dias_habiles_restantes-1} días hábiles"
                elif horas_restantes > 0:
                    dias_restantes = f"Te quedan {horas_restantes} horas y {minutos_restantes} minutos"
                else:
                    dias_restantes = f"Te quedan {minutos_restantes} minutos"
        else:
            dias_restantes = "Sin fecha límite"

        # Agregar los datos al listado de solicitudes
        solicitudes_data.append({
            'solicitud': solicitud,
            'numero_designios': numero_designios,
            'dias_restantes': dias_restantes
        })
    insumos = Insumo.objects.all()

    data = {
        'OPCIONES': OPCIONES,
        'Solicitudes': solicitudes_data,  # Lista con la información adicional
        'Usuarios': usuarios,
        'insumos': insumos,  # Lista de usuarios
    }

    return render(request, 'solicitude_llegadas.html', data,)

def calcular_dias_habiles(fecha_inicio, fecha_fin):
    dias_habiles = 0
    dia_actual = fecha_inicio

    while dia_actual <= fecha_fin:
        # Si el día actual no es sábado (5) ni domingo (6), lo contamos
        if dia_actual.weekday() < 5:  # 0 = Lunes, ..., 4 = Viernes
            dias_habiles += 1
        dia_actual += timedelta(days=1)
    
    return dias_habiles

# Relacionar Funcionario con User basado en nombres y apellidos
def obtener_usuario_por_funcionario(funcionario):
    return User.objects.filter(
        first_name=funcionario.nombre, last_name=funcionario.apellido
    ).first()

@login_required(login_url='/login/')
def control(request):
    # Obtener todas las solicitudes y tareas
    total_solicitudes = ProtocoloSolicitud.objects.count()
    total_tareas = Tarea.objects.filter(completada=True).count()

    # Contar solicitudes por estado
    estados = ProtocoloSolicitud.objects.values("estado").annotate(total=Count("estado"))
    estado_counts = {estado["estado"]: estado["total"] for estado in estados}

    # Definir estados posibles
    estados_posibles = ["RECIBIDO", "EN PROCESO", "EJECUTADO", "RECHAZADO"]
    for estado in estados_posibles:
        estado_counts.setdefault(estado, 0)

    # Trabajo Propio (Puntaje y Total de Solicitudes)
    profesionales_solicitudes = (
        ProtocoloSolicitud.objects.filter(profesional__isnull=False)
        .values("profesional__first_name", "profesional__last_name", "profesional_id")
        .annotate(
            trabajo_propio=Sum("valor_de_trabajo_funcionario"),
            total_solicitudes=Count("id")
        )
    )

    # Trabajo de Apoyo (Puntaje y Total de Solicitudes)
    profesionales_apoyos = (
        Apoyo_Protocolo.objects.filter(profesional__isnull=False)
        .values("profesional__first_name", "profesional__last_name", "profesional_id")
        .annotate(
            trabajo_apoyo=Sum("valor_de_trabajo"),
            total_solicitudes_apoyo=Count("id")
        )
    )

    # Tareas Internas (Filtrar solo tareas completadas)
    tareas_completadas = (
        Tarea.objects.filter(completada=True)
        .values("funcionario__nombre", "funcionario__apellido", "funcionario_id")
        .annotate(total_tareas=Count("id"))
    )

    tareas_con_apoyo = (
        Tarea.objects.filter(completada=True, apoyo__isnull=False)
        .values("funcionario_id")
        .annotate(total_apoyos=Count("id"))
    )
    # Combinar datos
    puntajes_por_profesional = {}

    # Agregar ProtocoloSolicitud
    for item in profesionales_solicitudes:
        full_name = f"{item['profesional__first_name']} {item['profesional__last_name']}"
        puntajes_por_profesional[item["profesional_id"]] = {
            "name": full_name,
            "trabajo_propio": float(item["trabajo_propio"] or 0),
            "trabajo_apoyo": 0,
            "total_solicitudes": item["total_solicitudes"],
            "total_solicitudes_apoyo": 0,
            "total_tareas": 0  # Inicialmente 0
        }
    
    solicitudes_per_profesional = (
        ProtocoloSolicitud.objects.filter(profesional__isnull=False)
        .values("profesional__first_name", "profesional__last_name")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    labels = [
    f"{item['profesional__first_name']} {item['profesional__last_name']}"
    for item in solicitudes_per_profesional
    ]

    # Agregar Apoyo_Protocolo
    for item in profesionales_apoyos:
        full_name = f"{item['profesional__first_name']} {item['profesional__last_name']}"
        if item["profesional_id"] in puntajes_por_profesional:
            puntajes_por_profesional[item["profesional_id"]]["trabajo_apoyo"] = float(item["trabajo_apoyo"] or 0)
            puntajes_por_profesional[item["profesional_id"]]["total_solicitudes_apoyo"] = item["total_solicitudes_apoyo"]
        else:
            puntajes_por_profesional[item["profesional_id"]] = {
                "name": full_name,
                "trabajo_propio": 0,
                "trabajo_apoyo": float(item["trabajo_apoyo"] or 0),
                "total_solicitudes": 0,
                "total_solicitudes_apoyo": item["total_solicitudes_apoyo"],
                "total_tareas": 0
            }

    # Agregar Tareas Internas
    for item in tareas_completadas:
        funcionario = Funcionario.objects.get(id=item["funcionario_id"])
        user = obtener_usuario_por_funcionario(funcionario)
        
        if user:
            if user.id in puntajes_por_profesional:
                puntajes_por_profesional[user.id]["total_tareas"] = item["total_tareas"]
            else:
                full_name = f"{funcionario.nombre} {funcionario.apellido}"
                puntajes_por_profesional[user.id] = {
                    "name": full_name,
                    "trabajo_propio": 0,
                    "trabajo_apoyo": 0,
                    "total_solicitudes": 0,
                    "total_solicitudes_apoyo": 0,
                    "total_tareas": item["total_tareas"]
                }

    for item in tareas_con_apoyo:
        funcionario = Funcionario.objects.get(id=item["funcionario_id"])
        user = obtener_usuario_por_funcionario(funcionario)

        if user:
            if user.id in puntajes_por_profesional:
                puntajes_por_profesional[user.id]["total_apoyos_tareas"] = item["total_apoyos"]
            else:
                full_name = f"{funcionario.nombre} {funcionario.apellido}"
                puntajes_por_profesional[user.id] = {
                    "name": full_name,
                    "trabajo_propio": 0,
                    "trabajo_apoyo": 0,
                    "total_solicitudes": 0,
                    "total_solicitudes_apoyo": 0,
                    "total_tareas": 0,
                    "total_apoyos_tareas": item["total_apoyos"]
                }



    # Ordenar por total de trabajo
    sorted_puntajes = sorted(
        puntajes_por_profesional.values(),
        key=lambda x: x["trabajo_propio"] + x["trabajo_apoyo"] + x["total_tareas"] + x.get("total_apoyos_tareas", 0),
        reverse=True
    )

    # Listas para gráficos
    labels = [item["name"] for item in sorted_puntajes]
    trabajo_propio_data = [item["trabajo_propio"] for item in sorted_puntajes]
    trabajo_apoyo_data = [item["trabajo_apoyo"] for item in sorted_puntajes]
    total_solicitudes_data = [item["total_solicitudes"] for item in sorted_puntajes]
    total_solicitudes_apoyo_data = [item["total_solicitudes_apoyo"] for item in sorted_puntajes]
    total_tareas_data = [item["total_tareas"] for item in sorted_puntajes]
    total_apoyos_tareas_data = [item.get("total_apoyos_tareas", 0) for item in sorted_puntajes]

    # Convertir a JSON
    labels_json = json.dumps(labels)
    trabajo_propio_json = json.dumps(trabajo_propio_data)
    trabajo_apoyo_json = json.dumps(trabajo_apoyo_data)
    total_solicitudes_json = json.dumps(total_solicitudes_data)
    total_solicitudes_apoyo_json = json.dumps(total_solicitudes_apoyo_data)
    total_tareas_json = json.dumps(total_tareas_data)
    total_apoyos_tareas_json = json.dumps(total_apoyos_tareas_data)


    context = {
        "total_solicitudes": total_solicitudes,
        "total_tareas": total_tareas,
        "labels_json": labels_json,
        "trabajo_porcentual_apoyo_json": trabajo_apoyo_json,
        "trabajo_porcentual_propio_json": trabajo_propio_json,
        "total_unitario_solicitudes_json": total_solicitudes_json,
        "total_unitario_solicitudes_apoyo_json": total_solicitudes_apoyo_json,
        "total_tareas_json": total_tareas_json,
        "total_apoyos_tareas_json": total_apoyos_tareas_json,
        "en_proceso": estado_counts.get("EN PROCESO", 0),
        "ejecutado": estado_counts.get("EJECUTADO", 0),
        "rechazado": estado_counts.get("RECHAZADO", 0),
    }

    return render(request, "Control.html", context)

def cambiar_contraseña(request):
    if request.method == 'POST':
            password1 = request.POST['password1']
            password2 = request.POST['password2']

            if password1 == password2:
                user = User.objects.get(id=request.user.id)
                user.set_password(password1)
                user.save()

                # Mantener al usuario autenticado después del cambio de contraseña
                update_session_auth_hash(request, user)

                return redirect('control')
            else:
                messages.error(request, 'Las contraseñas no coinciden. Inténtalo de nuevo.')

    return render(request, 'cambiar_contraseña.html')

def logout(request):
    auth_logout(request)
    return redirect('core_login')

def Gestion_imagen(request):
    if request.method == 'POST':
        form = ImagenForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('Gestion_imagen')  # Redirige a la misma vista después de guardar la imagen
    else:
        form = ImagenForm()

    # Obtener todas las imágenes
    imagenes = Imagen_sig.objects.all()
    paginator = Paginator(imagenes, 6)  # Muestra 6 imágenes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'Gestion_imagen.html', {'form': form, 'imagenes': imagenes,'page_obj': page_obj})


def eliminar_imagen(request, imagen_id):
    imagen = get_object_or_404(Imagen_sig, pk=imagen_id)
    
    if request.method == 'POST':
        # Guarda la ruta del archivo de la imagen
        ruta_archivo = imagen.archivo_adjunto.path
        # Elimina la imagen de la base de datos
        imagen.delete()
        # Elimina el archivo de la imagen del sistema de archivos
        if os.path.exists(ruta_archivo):
            os.remove(ruta_archivo)
        return redirect('Gestion_imagen')  # Redirige a la vista de gestión de imágenes después de eliminar la imagen
    
    return redirect('Gestion_imagen')  # Redirige a la vista de gestión de imágenes si la solicitud no es de tipo POST

def Gestion_pdf(request):
    if request.method == 'POST':
        form = PDFForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('Gestion_pdf')  # Redirige a la misma vista después de guardar la imagen
    else:
        form = PDFForm()

    # Obtener todas las imágenes
    pdf = PDF_sig.objects.all()
    paginator = Paginator(pdf, 6)  # Muestra 6 imágenes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'Gestion_PDF.html', {'form': form, 'PDF': pdf,'page_obj': page_obj})




def solicitudes_json(request):
    try:
        draw = int(request.POST.get("draw", 1))
        start = int(request.POST.get("start", 0))
        length = int(request.POST.get("length", 10))
        search_value = request.POST.get("search[value]", "")


        # ✅ Ordenamiento dinámico según DataTables
        order_column_index = int(request.GET.get("order[0][column]", 0))
        order_dir = request.GET.get("order[0][dir]", "asc")

        # Columnas que DataTables puede ordenar
        columns = [
            "id",                    # 0
            "orden_trabajo",        # 1
            "nombre_solicitante",   # 2
            "fecha",                # 3
            "departamento",
            "",
            "",
            "",
            "",
            "fecha_D",          # 4
            "fecha_T",          # 5

                                                                  # 4
        ]

        if order_column_index >= len(columns):
            order_column_index = 0

        order_column = columns[order_column_index]
        if order_column:
            if order_dir == "desc":
                order_column = f"-{order_column}"
        else:
            order_column = "-id"  # fallback

        # 🔐 Filtrar por usuario
        if request.user.is_superuser:
            solicitudes = ProtocoloSolicitud.objects.all()
        else:
            solicitudes = ProtocoloSolicitud.objects.filter(
                Q(profesional=request.user) | Q(solicitud__profesional=request.user)
            ).distinct()

        total_records = solicitudes.count()

        # 🔍 Filtro por búsqueda
        if search_value:
            solicitudes = solicitudes.filter(
                Q(nombre_solicitante__icontains=search_value) |
                Q(nombre_proyecto__icontains=search_value) |
                Q(departamento__icontains=search_value) |
                Q(codigo__icontains=search_value)
            )

        total_filtered = solicitudes.count()

        # 🔁 Aplicar ordenamiento y paginar
        solicitudes = solicitudes.order_by(order_column)
        paginator = Paginator(solicitudes, length)
        page_number = (start // length) + 1
        page = paginator.get_page(page_number)

        # 🧱 Armar estructura de datos
        data = []
        for solicitud in page:
            numero_designios = Registro_designio.objects.filter(protocolo=solicitud).count()

            apoyos = Apoyo_Protocolo.objects.filter(protocolo=solicitud)
            apoyos_lista = [
                {
                    "id": apoyo.profesional.id,
                    "nombre": f"{apoyo.profesional.first_name} {apoyo.profesional.last_name}",
                    "correo": apoyo.profesional.email,
                }
                for apoyo in apoyos
            ]

            # 📅 Días restantes
            if solicitud.estado == "RECHAZADO":
                dias_restantes = "Rechazado"
            elif solicitud.fecha_T:
                dias_restantes = "Trabajo terminado"
            elif solicitud.fecha_L:
                total_segundos = (solicitud.fecha_L - now()).total_seconds()
                if total_segundos < 0:
                    dias_pasados = abs(int(total_segundos // (24 * 3600)))
                    dias_restantes = f"Pasada por {dias_pasados} días"
                else:
                    dias_restantes = "Dentro del plazo"
            else:
                dias_restantes = "Sin fecha límite"

            archivos_adjuntos = ArchivoProtocolo.objects.filter(protocolo=solicitud)
            archivos_adjuntos_urls = [archivo.archivo.url for archivo in archivos_adjuntos] if archivos_adjuntos.exists() else []

            data.append({
                'id': solicitud.id,
                'departamento': solicitud.departamento,
                'direccion': solicitud.direccion,
                'nombre_solicitante': solicitud.nombre_solicitante,
                'nombre_proyecto': solicitud.nombre_proyecto,
                'corre_solicitante': solicitud.corre_solicitante,
                'area': solicitud.area,
                'objetivos': solicitud.objetivos,
                'cambios_posible': solicitud.cambios_posible,
                'fecha': solicitud.fecha.strftime('%Y-%m-%d %H:%M:%S') if solicitud.fecha else None,
                'codigo': solicitud.codigo,
                'orden_trabajo': solicitud.orden_trabajo,
                'fecha_D': solicitud.fecha_D.strftime('%Y-%m-%d %H:%M:%S') if solicitud.fecha_D else "Sin Fecha",
                'fecha_T': solicitud.fecha_T.strftime('%Y-%m-%d %H:%M:%S') if solicitud.fecha_T else "Sin Fecha",
                'fecha_L': solicitud.fecha_L.strftime('%Y-%m-%d') if solicitud.fecha_L else "Sin Fecha",
                'profesional_id': solicitud.profesional.id if solicitud.profesional else None,
                'profesional_nombre': f"{solicitud.profesional.first_name} {solicitud.profesional.last_name}" if solicitud.profesional else "Sin asignar",
                'profesional_correo': solicitud.profesional.email if solicitud.profesional else "Sin asignar",
                'tipo_limite': solicitud.tipo_limite,
                'estado': solicitud.estado,
                'enviado_correo': solicitud.enviado_correo,
                'enviado_correo_t': solicitud.enviado_correo_t,
                'numero_designios': numero_designios,
                'dias_restantes': dias_restantes,
                'archivos_adjuntos_urls': archivos_adjuntos_urls,
                'apoyos': apoyos_lista,
            })

        # 📤 Respuesta
        return JsonResponse({
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": total_filtered,
            "data": data,
            "id_user": request.user.id,
            "es_superuser": request.user.is_superuser,
        }, safe=False)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
